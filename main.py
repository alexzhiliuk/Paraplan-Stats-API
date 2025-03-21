from pprint import pprint

import requests
import json
from typing import Literal
from datetime import date, timedelta
import calendar
import openpyxl
import sys
import os
import logging
from dotenv import load_dotenv

from exceptions import AuthError, CsrfTokenError
from bot import send_report_to_tg

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, filename="logs.log",
                    format="%(asctime)s::%(levelname)s::%(message)s")

load_dotenv()


class ParaplanAPI:
    BASE_URL = "https://paraplancrm.ru/"
    USER_URL = BASE_URL + "api/open/user"
    LOGIN_URL = BASE_URL + "api/public/login"
    STUDENTS_URL = BASE_URL + "api/open/students/min-info"
    ATTENDANCES_STATUSES_URL = BASE_URL + "api/open/attendances/students/statuses"
    ATTENDANCES_URL_TEMPLATE = BASE_URL + "api/open/company/attendances/breakdown/group?date.year={year}&date.month={month}&date.day={day}&scheduleBreakdownAccessTypeSet=ATTENDANCES&scheduleBreakdownAccessTypeSet=LESSONS&scheduleBreakdownAccessTypeSet=PREBOOKINGS&scheduleBreakdownAccessTypeSet=SCHEDULE_MODIFICATIONS"
    IND_ATTENDANCES_URL_TEMPLATE = BASE_URL + "api/open/company/attendances/breakdown/individual?date.year={year}&date.month={month}&date.day={day}&scheduleBreakdownAccessTypeSet=ATTENDANCES&scheduleBreakdownAccessTypeSet=LESSONS&scheduleBreakdownAccessTypeSet=PREBOOKINGS&scheduleBreakdownAccessTypeSet=SCHEDULE_MODIFICATIONS"
    ATTENDANCES_FOR_SCREEN_URL_TEMPLATE = BASE_URL + "api/open/company/attendances/{attendance_id}/forAttendanceScreen"
    STUDENT_SUBSCRIPTIONS_URL_TEMPLATE = BASE_URL + "/api/open/students/{student_id}/subscriptions/paginated?page=1&size=10"
    GROUP_INFO_URL_TEMPLATE = BASE_URL + "api/open/groups/{group_id}"
    STUDENT_CARD_URL_TEMPLATE = "https://paraplancrm.ru/crm/#/students/{student_id}/groups"

    STATUSES = {
        "ATTENDED_TRIAL": ["fea4db4a-b812-a27f-1d02-998fc23f76b3", "78e0eab0-b4d4-9cd2-3c9a-bc862db3bbbc"]
    }

    LOGIN_DATA = json.dumps({
        "username": os.getenv("LOGIN", None),
        "password": os.getenv("PASS", None),
        "locale": "RU",
        "loginType": "KIDS_APP",
        "rememberMe": False,
        "captcha": ""
    })
    STUDENTS_DATA = json.dumps({
        "currentOnly": False
    })

    HEADERS = {
        'Content-Type': 'application/json',
    }

    def __init__(self):
        self.session = requests.Session()
        self.session.request("POST", self.LOGIN_URL, headers=self.HEADERS, data=self.LOGIN_DATA)

        if self.session.get(self.USER_URL).status_code != 200:
            raise AuthError("Не удалось залогиниться")

        csrf_token = self._get_csrf_token()
        if not csrf_token:
            raise CsrfTokenError("Не удалось получить CSRF токен")

        self._update_csrf_in_headers(csrf_token)

        self.current_month_period = self._get_month_period("current")
        self.previous_month_period = self._get_month_period("previous")
        self.next_month_period = self._get_month_period("next")
        self.current_week_period = self._get_current_week_period()
        self.after_current_week_period = self._get_period_after_current_week()

    def _get_csrf_token(self):
        return self.session.cookies.get("XSRF-TOKEN")

    def _update_csrf_in_headers(self, csrf_token):
        self.HEADERS["X-XSRF-TOKEN"] = csrf_token

    @staticmethod
    def _get_month_period(period: Literal["current", "previous", "next"]) -> tuple[date, date]:

        today = date.today()
        period_end_date = today - timedelta(days=today.day)

        if period == "current":
            pass
        elif period == "previous":
            period_end_date = period_end_date - timedelta(days=period_end_date.day)
        elif period == "next":
            period_start_date = period_end_date + timedelta(days=1)
            period_last_day_num = calendar.monthrange(period_start_date.year, period_start_date.month)[1]
            period_end_date = date(period_start_date.year, period_start_date.month, period_last_day_num)

        return period_end_date.replace(day=1), period_end_date

    @staticmethod
    def _get_current_week_period() -> tuple[date, date]:
        today = date.today()
        week_start_date = today - timedelta(days=today.weekday())
        week_end_date = week_start_date + timedelta(days=6)

        return week_start_date, week_end_date

    @staticmethod
    def _get_period_after_current_week() -> tuple[date, None]:
        today = date.today()
        week_start_date = today - timedelta(days=today.weekday())
        next_week_start_date = week_start_date + timedelta(days=7)
        return next_week_start_date, None

    @staticmethod
    def _format_subs_end_date(end_date: dict):
        months = {1: "Января", 2: "Февраля", 3: "Марта", 4: "Апреля", 5: "Мая", 6: "Июня",
                  7: "Июля", 8: "Августа", 9: "Сентября", 10: "Октября", 11: "Ноября", 12: "Декабря"}
        return f"{end_date['day']} {months[end_date['month']]} {end_date['year']}"

    @staticmethod
    def _convert_subs_end_date_to_date(end_date: dict):
        return date(end_date["year"], end_date["month"], end_date["day"])

    @staticmethod
    def _get_start_period_parameters(start_date: date):
        return f"&from.day={start_date.day}&from.month={start_date.month}&from.year={start_date.year}"

    @staticmethod
    def _get_end_period_parameters(end_date: date):
        return f"&to.day={end_date.day}&to.month={end_date.month}&to.year={end_date.year}"

    def get_attendances_statuses(self):

        return self.session.get(self.ATTENDANCES_STATUSES_URL).json()

    def _get_students(self) -> list:
        response = self.session.post(self.STUDENTS_URL, headers=self.HEADERS, data=self.STUDENTS_DATA)
        return response.json()["studentList"]

    def _get_student_subscriptions(self, student_id: str, period: tuple[date | None, date | None] = None,
                                   filter_by_end_date: bool = False) -> list:
        url = self.STUDENT_SUBSCRIPTIONS_URL_TEMPLATE.format(student_id=student_id)
        if period:
            if period[0]:
                url += self._get_start_period_parameters(period[0])
            if period[1]:
                url += self._get_end_period_parameters(period[1])
        subscriptions = self.session.get(url).json().get("itemList")
        subscriptions = list(filter(lambda item: item["lessonQuantity"] > 1 and item["endDate"], subscriptions))

        if filter_by_end_date and period:
            subscriptions = self._filter_subscriptions_by_end_date(subscriptions, period)

        return subscriptions

    def _get_group_info(self, group_id):
        return self.session.get(self.GROUP_INFO_URL_TEMPLATE.format(group_id=group_id)).json().get("group")

    def _filter_subscriptions_by_end_date(self, subscriptions: list, period: tuple[date | None, date | None]) -> list:
        if period[0] and period[1]:
            return list(filter(
                lambda item: period[0] <= self._convert_subs_end_date_to_date(item["endDate"]) <= period[1],
                subscriptions
            ))
        if period[0]:
            return list(filter(
                lambda item: period[0] <= self._convert_subs_end_date_to_date(item["endDate"]),
                subscriptions
            ))
        if period[1]:
            return list(filter(
                lambda item: self._convert_subs_end_date_to_date(item["endDate"]) <= period[1],
                subscriptions
            ))

    def _is_student_has_non_renewed_subs_in_month(self, student) -> tuple[bool, str | None]:
        previous_period_subs = self._get_student_subscriptions(student["id"], self.previous_month_period)
        if not previous_period_subs:
            return False, None

        current_period_subs = self._get_student_subscriptions(student["id"], self.current_month_period)
        if current_period_subs:
            return False, None

        return True, self._format_subs_end_date(previous_period_subs[0]["endDate"])

    def _get_attendances_ids(self, attendance_date: date) -> list:
        attendance_list = self.session.get(
            self.ATTENDANCES_URL_TEMPLATE.format(year=attendance_date.year, month=attendance_date.month,
                                                 day=attendance_date.day)
        ).json()["breakdown"]["attendanceList"]

        return [attendance.get("id") for attendance in attendance_list]

    def _get_filtered_attendees(self, attendance) -> [list, str, str]:

        attendees_list = [{"id": attendee["studentInfo"]["id"], "name": attendee["studentInfo"]["name"]} for
                          attendee in attendance["attendeeList"] if
                          attendee["statusId"] in self.STATUSES["ATTENDED_TRIAL"]]

        attendance_time = f"{attendance['dateTime']['hour']}:{str(attendance['dateTime']['minute']).zfill(2)}"
        attendance_teachers = " ".join(
            [teacher["teacherInfo"]["name"] for teacher in attendance["teacherList"]])

        return attendees_list, attendance_time, attendance_teachers

    def _get_students_attended_individual_trial(self, attendance_date: date) -> list:
        group_list = self.session.get(self.IND_ATTENDANCES_URL_TEMPLATE.format(year=attendance_date.year,
                                                                               month=attendance_date.month,
                                                                               day=attendance_date.day)).json()[
            "groupList"]
        students_attended_trial_and_has_subscription = []

        for group in group_list:
            for attendance in group.get("attendanceList", []):
                # Получаем учеников только с нужным статусом
                attendees_list, attendance_time, attendance_teachers = self._get_filtered_attendees(attendance)

                # Проверка наличия подписки у студента
                students_attended_trial_and_has_subscription += self._get_row_data_for_student_attended_trial(
                    attendees_list, attendance_date, attendance_time, attendance_teachers)

        return students_attended_trial_and_has_subscription

    def _get_students_attended_group_trial(self, attendance_date: date) -> list:
        attendances_ids = self._get_attendances_ids(attendance_date)
        students_attended_trial_and_has_subscription = []

        # Перебор всех занятий для поиска учеников
        for attendance_id in attendances_ids:
            attendance = self.session.get(
                self.ATTENDANCES_FOR_SCREEN_URL_TEMPLATE.format(attendance_id=attendance_id)).json()["attendance"]

            # Получаем учеников только с нужным статусом
            attendees_list, attendance_time, attendance_teachers = self._get_filtered_attendees(attendance)

            # Проверка наличия подписки у студента
            students_attended_trial_and_has_subscription += self._get_row_data_for_student_attended_trial(
                attendees_list, attendance_date, attendance_time, attendance_teachers)

        return students_attended_trial_and_has_subscription

    def _get_row_data_for_student_attended_trial(self, attendees_list: list, attendance_date: date,
                                                 attendance_time: str, attendance_teachers: str) -> list:
        students_attended_trial_and_has_subscription = []
        for attendee in attendees_list:
            subscriptions = self._get_student_subscriptions(attendee["id"], (attendance_date, None))
            is_subscribed = "Куплен" if len(subscriptions) else "Не куплен"

            students_attended_trial_and_has_subscription.append(
                {
                    "name": attendee["name"],
                    "link": self.STUDENT_CARD_URL_TEMPLATE.format(student_id=attendee["id"]),
                    "date": f"{attendance_date} {attendance_time}",
                    "is_subscribed": is_subscribed,
                    "teachers": attendance_teachers
                }
            )
            logger.info(f"Student {attendee['id']} processed")
        return students_attended_trial_and_has_subscription

    def get_students_with_non_renewed_subscription_in_month(self) -> list:
        students_with_non_renewed_subscription = []
        students = self._get_students()

        for student in students:
            _, subs_end_date = self._is_student_has_non_renewed_subs_in_month(student)
            if not subs_end_date:
                continue

            group_info = self._get_group_info(self._get_student_subscriptions(student["id"])[0]["groupList"][0]["id"])
            students_with_non_renewed_subscription.append(
                {
                    "name": student["name"],
                    "link": self.STUDENT_CARD_URL_TEMPLATE.format(student_id=student["id"]),
                    "subs_end_date": subs_end_date,
                    "type": group_info.get("type"),
                    "teacher": group_info["teacherList"][0]["name"] if group_info["teacherList"] else "-"
                }
            )
            logger.info(f"Student {student['id']} processed")

        return students_with_non_renewed_subscription

    def get_students_week_subscriptions_info(self) -> dict:
        students_ids_who_have_non_renewed_subscription = []
        students_ids_who_renewed_subscription = []
        students = self._get_students()

        for student in students:

            current_week_subscriptions = self._get_student_subscriptions(student["id"], self.current_week_period,
                                                                         filter_by_end_date=True)
            if current_week_subscriptions:

                after_current_week_subscriptions = self._get_student_subscriptions(student["id"],
                                                                                   self.after_current_week_period,
                                                                                   filter_by_end_date=True)
                if after_current_week_subscriptions:

                    group_info = self._get_group_info(after_current_week_subscriptions[0]["groupList"][0]["id"])
                    students_ids_who_renewed_subscription.append({
                        "link": self.STUDENT_CARD_URL_TEMPLATE.format(student_id=student["id"]),
                        "type": group_info.get("type"),
                        "teacher": group_info["teacherList"][0]["name"] if group_info["teacherList"] else "-"
                    })

                else:
                    group_info = self._get_group_info(current_week_subscriptions[0]["groupList"][0]["id"])
                    students_ids_who_have_non_renewed_subscription.append({
                        "link": self.STUDENT_CARD_URL_TEMPLATE.format(student_id=student["id"]),
                        "type": group_info.get("type"),
                        "teacher": group_info["teacherList"][0]["name"] if group_info["teacherList"] else "-"
                    })

                logger.info(f"Student {student['id']} processed")

        return {
            "have_non_renewed_subscription": students_ids_who_have_non_renewed_subscription,
            "who_renewed_subscription": students_ids_who_renewed_subscription
        }

    def get_students_with_ending_subscription_in_next_month(self) -> list:
        students_with_ending_subscription_in_next_month = []
        students = self._get_students()

        for student in students:
            subscriptions_ending_in_next_month = self._get_student_subscriptions(
                student["id"],
                self._get_month_period("next"),
                filter_by_end_date=True
            )

            for subscription in subscriptions_ending_in_next_month:
                students_with_ending_subscription_in_next_month.append({
                    "name": student["name"],
                    "link": self.STUDENT_CARD_URL_TEMPLATE.format(student_id=student["id"]),
                    "subs_end_date": self._format_subs_end_date(subscription["endDate"]),
                    "total_price": subscription["totalPrice"]
                })

                logger.info(f"Student {student['id']} processed")

        return students_with_ending_subscription_in_next_month

    def get_students_attended_trial(self, period: tuple[date, date]) -> list:

        start_date, end_date = period
        students_attended_trial_and_has_subscription = []

        # Перебор всех дней для сбора занятий
        current_date = start_date
        while current_date <= end_date:
            students_attended_trial_and_has_subscription += self._get_students_attended_group_trial(current_date)
            students_attended_trial_and_has_subscription += self._get_students_attended_individual_trial(current_date)

            current_date += timedelta(days=1)

        return students_attended_trial_and_has_subscription

    def create_excel_file_students_with_non_renewed_subscription_in_month(self, filename) -> None:

        students = self.get_students_with_non_renewed_subscription_in_month()

        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]

        ws["A1"] = "Имя ученика"
        ws["B1"] = "Дата окончания абонемента"
        ws["C1"] = "Ссылка на карточку ученика"
        ws["D1"] = "Статус занятия"
        ws["E1"] = "Преподаватель"

        for row_index, student in enumerate(students, start=2):
            ws[f"A{row_index}"] = student["name"]
            ws[f"B{row_index}"] = student["subs_end_date"]
            ws[f"C{row_index}"] = student["link"]
            ws[f"D{row_index}"] = student["type"]
            ws[f"E{row_index}"] = student["teacher"]

        wb.save(filename=filename)
        logger.info("Excel file with non-renewed subs in month was created")

    def create_excel_file_with_students_week_subscriptions_info(self, filename) -> None:

        subs_info = self.get_students_week_subscriptions_info()

        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]

        ws["A1"] = "Всего"
        ws["A2"] = f"{len(subs_info["have_non_renewed_subscription"]) + len(subs_info["who_renewed_subscription"])}"
        ws["B1"] = "Непродлившие"
        ws["B2"] = f"{len(subs_info["have_non_renewed_subscription"])}"
        ws["C1"] = "Статус занятия"
        ws["D1"] = "Преподаватель"
        ws["E1"] = "Продлившие"
        ws["E2"] = f"{len(subs_info["who_renewed_subscription"])}"
        ws["F1"] = "Статус занятия"
        ws["G1"] = "Преподаватель"

        for row_index, student_link in enumerate(subs_info["have_non_renewed_subscription"], start=3):
            ws[f"B{row_index}"] = student_link["link"]
            ws[f"C{row_index}"] = student_link["type"]
            ws[f"D{row_index}"] = student_link["teacher"]

        for row_index, student_link in enumerate(subs_info["who_renewed_subscription"], start=3):
            ws[f"E{row_index}"] = student_link["link"]
            ws[f"F{row_index}"] = student_link["type"]
            ws[f"G{row_index}"] = student_link["teacher"]

        wb.save(filename=filename)
        logger.info("Excel file with students week subs info was created")

    def create_excel_students_with_ending_subscription_in_next_month(self, filename: str) -> None:

        students = self.get_students_with_ending_subscription_in_next_month()

        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]

        ws["A1"] = "Имя"
        ws["B1"] = "Сумма"
        ws["C1"] = "Дата окончания абонемента"
        ws["D1"] = "Ссылка на карточку ученика"

        for row_index, student in enumerate(students, start=2):
            ws[f"A{row_index}"] = student["name"]
            ws[f"B{row_index}"] = student["total_price"]
            ws[f"C{row_index}"] = student["subs_end_date"]
            ws[f"D{row_index}"] = student["link"]

        wb.save(filename=filename)
        logger.info("Excel file with students ending subs in next month was created")

    def create_excel_students_attended_trial(self, filename: str, period: tuple[date, date]) -> None:

        students = self.get_students_attended_trial(period)

        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]

        ws["A1"] = "Имя ученика"
        ws["B1"] = "Ссылка на карточку ученика"
        ws["C1"] = "Дата пробного занятия"
        ws["D1"] = "Статус абонемента"
        ws["E1"] = "Педагог"

        for row_index, student in enumerate(students, start=2):
            ws[f"A{row_index}"] = student["name"]
            ws[f"B{row_index}"] = student["link"]
            ws[f"C{row_index}"] = student["date"]
            ws[f"D{row_index}"] = student["is_subscribed"]
            ws[f"E{row_index}"] = student["teachers"]

        wb.save(filename=filename)
        logger.info("Excel file with students attended trial was created")


def test():
    paraplan = ParaplanAPI()
    pprint(list(
        filter(lambda status: "посетил" in status["name"].lower(), paraplan.get_attendances_statuses()["statusList"])))


def main():
    actions_list = ["current-month", "current-week", "next-month", "month-conversion-of-trial-sessions",
                    "week-conversion-of-trial-sessions"]

    if len(sys.argv) < 2:
        message = f"Не указан тип действия\nИспользуйте {' | '.join(actions_list)}"
        logger.error(message)
        print(message)
        return

    if sys.argv[1] not in actions_list:
        message = f"Используйте {' | '.join(actions_list)}"
        logger.error(message)
        print(message)
        return

    paraplan = ParaplanAPI()

    if sys.argv[1] == "month-conversion-of-trial-sessions":
        filename = "conversion-of-trial-sessions.xlsx"
        paraplan.create_excel_students_attended_trial(filename, paraplan.current_month_period)
        send_report_to_tg(filename)
    if sys.argv[1] == "week-conversion-of-trial-sessions":
        filename = "conversion-of-trial-sessions.xlsx"
        paraplan.create_excel_students_attended_trial(filename, paraplan.current_week_period)
        send_report_to_tg(filename)
    if sys.argv[1] == "current-month":
        filename = "students-month.xlsx"
        paraplan.create_excel_file_students_with_non_renewed_subscription_in_month(filename)
        send_report_to_tg(filename)
    if sys.argv[1] == "current-week":
        filename = "students-week-info.xlsx"
        paraplan.create_excel_file_with_students_week_subscriptions_info(filename)
        send_report_to_tg(filename)
    if sys.argv[1] == "next-month":
        filename = "students-predicts.xlsx"
        paraplan.create_excel_students_with_ending_subscription_in_next_month(filename)
        send_report_to_tg(filename)


if __name__ == "__main__":
    try:
        main()
    except AuthError as err:
        logger.error(err)
        print(err)
    except CsrfTokenError as err:
        logger.error(err)
        print(err)
    except Exception as err:
        logger.error(err, exc_info=True)
        print(err)
